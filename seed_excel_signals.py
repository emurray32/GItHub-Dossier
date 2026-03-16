"""One-time data migration: clear all existing signals and import from Excel.

Loads the 109 intent signals from the curated Excel spreadsheet,
replacing all existing signals, accounts, and prospects.

Usage:
    python seed_excel_signals.py           # Run import
    python seed_excel_signals.py --dry-run  # Preview without changing anything
"""

import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

# Initialize the database before importing services
from database import init_db
init_db()

from v2.services.ingestion_service import ingest_excel


EXCEL_FILE = os.path.join(
    os.path.dirname(__file__),
    'attached_assets',
    'Example_Intent_Signal_Sheet_(1)_1773688913655.xlsx',
)


def main():
    dry_run = '--dry-run' in sys.argv

    if not os.path.exists(EXCEL_FILE):
        print(f"Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    if dry_run:
        print("DRY RUN — reading file only, no changes.\n")
        # Just show what would be imported
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            non_empty = [r for r in rows[1:] if r[1]]  # Column B = Company
            print(f"Sheet '{sn}': {len(non_empty)} companies with data")
            if non_empty:
                print(f"  Headers: {rows[0][:9]}")
                for r in non_empty[:3]:
                    print(f"  - {r[1]}: {r[3]} / {r[4]}")
                if len(non_empty) > 3:
                    print(f"  ... and {len(non_empty) - 3} more")
        wb.close()
        print("\nRun without --dry-run to clear existing signals and import.")
        return

    print("Loading Excel file and clearing all existing signals...")

    with open(EXCEL_FILE, 'rb') as f:
        file_content = f.read()

    result = ingest_excel(
        file_content=file_content,
        source_label='seed_excel_import',
        created_by='seed_script',
        clear_existing=True,
    )

    totals = result.get('totals', {})
    cleared = totals.get('cleared', 0)
    created = totals.get('signals_created', 0)
    accounts = totals.get('accounts_created', 0)
    matched = totals.get('accounts_matched', 0)
    errors = totals.get('errors', [])
    sheets = result.get('sheets_processed', 0)

    print(f"\nCleared {cleared} old signals.")
    print(f"Processed {sheets} sheet(s).")
    print(f"Created {created} new signals from {accounts} new accounts ({matched} matched existing).")

    if errors:
        print(f"\n{len(errors)} error(s):")
        for e in errors[:10]:
            print(f"  - {e}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")
    else:
        print("\nNo errors.")

    enrichment = totals.get('enrichment') or (result.get('sheets', [{}])[0].get('enrichment') if result.get('sheets') else None)
    if enrichment:
        if enrichment.get('accounts_enriched'):
            print(f"Apollo enriched {enrichment['accounts_enriched']} accounts.")
        if enrichment.get('signals_evaluated'):
            print(f"BDR scored {enrichment['signals_evaluated']} signals.")

    print("\nDone.")


if __name__ == '__main__':
    main()
